import pandas as pd
import openpyxl
import re
import datetime


file = './sample2.xlsx'
writingFile = './Book.xlsx'
xls = pd.ExcelFile(file)

df =  pd.read_excel(xls, "Μηνιαίοι Λογαριασμοί", skiprows=3)
# wDfLoipa = wb["Λοιπα"]
# wDfEsoda = wb["Εσοδα"]
# wDfMetrita = wb["Μετρητα"]
# wDfYgeia = wb["Υγεία"]
# wDfSpiti = wb["Σπιτι"]
# wDfPsihagogia = wb["Ψυχαγωγια"]
# wDfYpoxrewseis = wb[ "Υποχρεωσεις"]
# wDfPsonia = wb[ "Ψωνια"]
# wDfMetakinisi = wb[ "Μετακινηση"]
def foo (a, writingFile, sheetname):
    # wb = openpyxl.load_workbook(writingFile)
    for date,price,desc in a:
        writiInXLS(writingFile, date,price,desc,sheetname)
        # wb.save(writingFile)
    # wb.save(writingFile)
    
def writiInXLS(writingFile, date, price, desc, sheetname):
    wb = openpyxl.load_workbook(writingFile)
    max_rows = (wb[sheetname].max_row)

    # max_rows = len(pd.read_excel(pd.ExcelFile(writingFile),  sheetname))
    print("max rows: ",max_rows)
    if(max_rows==0):
        max_rows=1
    _ = wb[sheetname].cell(row = max_rows+1, column = 1, value=date)
    _ = wb[sheetname].cell(row = max_rows+1, column = 2, value=price)
    _ = wb[sheetname].cell(row = max_rows+1, column = 3, value=desc)
    wb.save(writingFile)
    



i=0
mydict = {"Ψυχαγωγια":[], "Ψωνια":[], "Σπιτι":[], "Υγεια":[], "Μετακινηση":[], "Υποχρεωσεις":[], "Λοιπα":[], "Εσοδα":[], "GROUP9":[]}
for y in range(len(df)):
    date = datetime.datetime.strptime(df.loc[y]["Ημ/νία Εγγραφής"], "%d/%m/%Y")
    price = df.loc[y]["Ποσό"] 
    desc =(df.loc[y]["Περιγραφή Συναλλαγής"])
    
    if(price<0): #----------------------------------------------------------------------------------------------------------------------------------esoda
       mydict["Εσοδα"].append((date,price,desc))
    #    print("[GROUP8] date: ", date, " desc: ", desc, " price: ", price) 
    elif(re.search("^PLAISIO|^COSMOTE|PLAISIO|OPUS | RETAIL WORLD|AMAZON|AMZ", desc)): #--------------------------------------------------------PSIXAGIA
        mydict["Ψυχαγωγια"].append((date,price,desc))  
        # print("[GROUP1] date: ", date, " desc: ", desc, " price: ", price)
    elif(re.search("^MARKS & SPENCER|SKROUTZ|HONDOSCENTER|eΦooΔ|TO DIAMANTI|JYSK|MAX STORES|PYXIDA|COFFEE BERRY", desc)): #-------------------------------------PSONIA
        mydict["Ψωνια"].append((date,price,desc))
        # print("[GROUP2] date: ", date, " desc: ", desc, " price: ", price) 
    elif(re.search("^AB|^SKLAVENITIS|^LIDL|^PET|^ARTOPOIIA|^DELENIKAS|^THEMART |MY MARKET|GALAKTOCOMICA|GALAXIAS|MYLONAS|KOLPOS KALONIS PSARAGO|O THOMAS", desc)): #-----SPITI
        mydict["Σπιτι"].append((date,price,desc))
        # writiInXLS(writingFile, date,price, desc, "Σπιτι")
        # max_rows = len(pd.read_excel(pd.ExcelFile(writingFile),  "Σπιτι"))
        # print("after max rows: ",max_rows)
        # i+=1
        
        # writingXls.save()
        # print("[GROUP3] date: ", date, " desc: ", desc, " price: ", price)
    elif(re.search("DOUZENI|PANAGIOTOU NIKO|BIOIATRIKI|FARMAKEIO|VIOIATRIKI|STAVROS SCHIZAS", desc)): #-----------------------------------------------------------YGEIA
        mydict["Υγεια"].append((date,price,desc))
        # print("[GROUP4] date: ", date, " desc: ", desc, " price: ", price)
    elif(re.search("^AVIN|^ATTIKI ODOS|MAKRAION|NEA ODOS| BP", desc)): #-------------------------------------------------------------------------------------METAKINHSH
        mydict["Μετακινηση"].append((date,price,desc))
        # print("[GROUP5] date: ", date, " desc: ", desc, " price: ", price)
    elif(re.search("^EYDAP|DEI", desc)): #------------------------------------------------------------------------------------------------------YPOXREWSEIS
        mydict["Υποχρεωσεις"].append((date,price,desc))
        # print("[GROUP6] date: ", date, " desc: ", desc, " price: ", price)
    # elif(re.search("^DOYZENI", desc)):
    #     mydict["GROUP4"]+=1 
    #     print("[GROUP4] date: ", date, " desc: ", desc, " price: ", price)
    # elif(re.search("^DOYZENI", desc)):
    #     mydict["GROUP4"]+=1 
    #     print("[GROUP4] date: ", date, " desc: ", desc, " price: ", price)
    else: #----------------------------------------------------------------------------------------------------------------------------------------LOIPA
        mydict["Λοιπα"].append((date,price,desc))
        print("[Λοιπα] date: ", date, " desc: ", desc, " price: ", price)
for group in mydict.keys():
    foo(mydict[group],writingFile,group)
print(mydict)

        