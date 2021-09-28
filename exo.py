

import pandas as pd
import openpyxl
import xlsxwriter
from pathlib import Path

xlsx_file = Path( '2020.xlsx')

wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active
sum=0

sum_per_table=[]
for row in sheet.iter_rows(7, 212):
    value=row[4].value
    
    sum=int(0 if value is None else value)+sum
    type_of_trans=row[0].value
  
print("sum for 2020",sum)
#print(table)
xlsx_file = Path( '2021.xlsx')

wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active
eksoda=0
esoda=0
tableofcategories={}
tableofcharges={}
valueoftranminus={}
valueoftranplus={}
print("--------------------------------------")
wb = openpyxl.Workbook()
possiblecharges=["ΑΡΓΥΡΩ ΑΓΓΕΛΗ", "AMZ", "PAYPAL", "BANGGOOD","Revolut" ]
for row in sheet.iter_rows(7, 212):
    value=row[4].value
    k=int(0 if value is None else value)
    type_of_trans=row[0].value
    top10charges=row[3].value
    
    for x in possiblecharges:
        if x not in top10charges:
            continue
        else:
            

            if x not in tableofcharges:
                tableofcharges[x]=1
                print ("i am here", row[4].value)
                if k <0:
                    valueoftranminus[x]=k
                else:
                    valueoftranplus[x]=k
            else :
                print("i am here2", k)
                print("looking for ", x, "in ",tableofcharges)
                tableofcharges[x]=tableofcharges[x]+1
                if k<0:
                    temp=valueoftranminus.get(x, 0)
                    valueoftranminus[x]=temp+ k
                    temp=0
                else:
                    temp=valueoftranplus.get(x, 0)
                    valueoftranplus[x]=temp+k
                    temp=0
    print("done", k)
    if type_of_trans not in tableofcategories:
        tableofcategories[type_of_trans]=1
    else:
        tableofcategories[type_of_trans]=tableofcategories[type_of_trans]+1
    
    if k>0:
        esoda=k+esoda
    else:
        eksoda=eksoda+k

wb = openpyxl.Workbook()

dest_filename = '1.xlsx'
ws3 = wb.create_sheet(title="Categories")
for j in tableofcharges:
    print(j,"-->",tableofcharges[j])
row = 1
col = 1
for y in tableofcategories:
    _ = ws3.cell(column=col, row=row, value=y)
    _ = ws3.cell(column=col+1, row=row, value=tableofcategories[y])
    row+=1
row=1
col=3
for m in tableofcharges:
    _ = ws3.cell(column=col, row=row, value=m)
    _ = ws3.cell(column=col+1, row=row, value=tableofcharges[m])
    _ = ws3.cell(column=col+2, row=row, value=valueoftranplus[m])
    temp=valueoftranminus.get(m,0)
    _ = ws3.cell(column=col+3, row=row, value=temp)
    row+=1
wb.save(filename = dest_filename)
print("esoda for 2021",esoda)
print("eksoda for 2021",eksoda)
print("sum for 2021", esoda+eksoda)